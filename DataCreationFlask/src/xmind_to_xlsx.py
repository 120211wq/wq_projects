import os
import zipfile

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from xmindparser import xmind_to_dict
import json


class xmind_to_xlsx():

    def load_xmind(self, filename, template):
        self.out = xmind_to_dict(filename)
        self.story = self.out[0]['topic']['topics']
        self.top_titile = self.out[0]['topic']['title']
        self.branch_code = self.get_branch_code(template)
        self.excel_public_name = filename.split('/')[-1].split('.')[0]+'_public.xlsx'
        self.excel_candao_name = filename.split('/')[-1].split('.')[0] + '_candao.xlsx'
        self.zip_name = filename.split('/')[-1].split('.')[0] + '.rar'
        self.creat_xlsx()

    def get_branch_code(self,template):
        branch = self.top_titile.split('/')[0].replace('/','')
        wb = openpyxl.load_workbook(template)
        sheet = wb['系统数据']
        for i in range(2, sheet.max_row + 1):
            bc = sheet.cell(row=i, column=5).value
            if branch in bc:
                branch_code = '('+ bc.split('(')[1]
                return branch_code
        else:
            return 0

    def get_model_code(self,model,template):
        wb = openpyxl.load_workbook(template)
        sheet = wb['系统数据']
        for i in range(2, sheet.max_row + 1):
            mo = sheet.cell(row=i, column=1).value
            if self.top_titile+'/'+model in mo:
                model_code = '('+ mo.split('(')[1]
                return model_code
        else:
            return 0


    def creat_xlsx(self):
        self.workbook_public = openpyxl.Workbook()
        self.workbook_candao = openpyxl.Workbook()
        self.workbook_public.save(os.getcwd()+'/src/excel_output/'+self.excel_public_name)
        self.workbook_candao.save(os.getcwd()+'/src/excel_output/'+self.excel_candao_name)
        self.workbook_public = openpyxl.load_workbook(os.getcwd()+'/src/excel_output/'+self.excel_public_name)
        self.workbook_candao = openpyxl.load_workbook(os.getcwd()+'/src/excel_output/'+self.excel_candao_name)
        self.ws_public = self.workbook_public['Sheet']
        self.row0_public = ["用例编号", '所属模块', '功能', '用例标题', '前置条件', '步骤', '预期', '优先级', '测试结果', '备注']
        for i in range(0, len(self.row0_public)):
            self.ws_public.cell(row=1, column=i + 1).value = self.row0_public[i]
        self.ws_candao = self.workbook_candao['Sheet']
        self.row0_candao = ["分支", '所属模块', '相关需求', '用例标题', '前置条件', '步骤', '预期', '关键词', '优先级', '用例类型',
                            '适用阶段','用例状态']
        for i in range(0, len(self.row0_candao)):
            self.ws_candao.cell(row=1, column=i + 1).value = self.row0_candao[i]
        self.workbook_public.save(os.getcwd()+'/src/excel_output/'+self.excel_public_name)
        self.workbook_candao.save(os.getcwd()+'/src/excel_output/'+self.excel_candao_name)

    def writeExcel_public(self, row, case):
        sort = 1
        for value in case:
            self.ws_public.cell(row=row, column=sort).value = value
            sort = sort + 1
            self.workbook_public.save(os.getcwd()+'/src/excel_output/'+self.excel_public_name)

    def writeExcel_candao(self, row, case):
        sort = 1
        for value in case:
            self.ws_candao.cell(row=row, column=sort).value = value
            sort = sort + 1
            self.workbook_candao.save(os.getcwd()+'/src/excel_output/'+self.excel_candao_name)

    def write_excel(self,template):
        self.rowNum = 2
        error = ''
        if len(self.story) == 0:
            print('无模块，请检查用例是否完整')
            return '无模块，请检查用例是否完整'
        else:
            for i in range(len(self.story)):
                model_leaf_title = self.story[i]['title']
                model_code = self.get_model_code(model_leaf_title,template)
                model_name = self.top_titile+'/'+model_leaf_title
                try:
                    model_leaf_topics = self.story[i]['topics']
                except KeyError:
                    model_leaf_topics = 0
                if model_leaf_topics == 0:
                    print('【' + model_leaf_title + '】' + '下无功能点，请检查用例是否完整')
                    error = error + '【' + model_leaf_title + '】' + '下无功能点，请检查用例是否完整'+'\n'
                    continue
                else:
                    for j in range(len(model_leaf_topics)):
                        func_title = model_leaf_topics[j]['title']
                        try:
                            func_topics = model_leaf_topics[j]['topics']
                        except KeyError:
                            func_topics = 0
                        if func_topics == 0:
                            print('【' + func_title + '】' + '模块下无用例，请检查用例是否完整')
                            error = error + '【' + func_title + '】' + '模块下无用例，请检查用例是否完整' + '\n'
                            continue
                        else:
                            for k in range(len(func_topics)):
                                case_title = func_topics[k]['title']
                                if ':' in case_title:
                                    priority = case_title.split(':')[1].split('_')[0][-1]
                                elif '：' in case_title:
                                    priority = case_title.split('：')[1].split('_')[0][-1]
                                else:
                                    priority = case_title.split('_')[1][-1]
                                case_name = case_title.split('_')[-1]
                                case_public = []
                                case_candao = []
                                setp = ''
                                except_result = ''
                                before = ''
                                try:
                                    case_topic = func_topics[k]['topics']
                                except KeyError:
                                    case_topic = 0
                                if case_topic == 0:
                                    print('【' + func_title + '】' + '用例下无具体步骤，请检查用例是否完整')
                                    error = error + '【' + func_title + '】' + '用例下无具体步骤，请检查用例是否完整' + '\n'
                                    continue
                                else:
                                    num = 1
                                    for l in range(len(case_topic)):
                                        if len(setp)>0:
                                            setp = setp + '\n'
                                        if len(except_result)>0:
                                            except_result = except_result + '\n'
                                        if 'before' in case_topic[l]['title']:
                                            if ':' in case_topic[l]['title']:
                                                before = case_topic[l]['title'].split(':')[1]
                                            elif '：' in case_topic[l]['title']:
                                                before = case_topic[l]['title'].split('：')[1]
                                            continue
                                        setp = setp + str(num)+'.'+case_topic[l]['title']
                                        try:
                                            except_result = except_result + str(num) + '.' + case_topic[l]['topics'][0][
                                                'title']
                                        except KeyError:
                                            pass
                                        num += 1
                                    case_public.append('')
                                    case_public.append(model_name)
                                    case_public.append(func_title)
                                    case_public.append(case_name)
                                    case_public.append(before)
                                    case_public.append(setp)
                                    case_public.append(except_result)
                                    case_public.append(priority)
                                    self.writeExcel_public(self.rowNum, case_public)
                                    case_candao.append(self.branch_code)
                                    case_candao.append(model_code)
                                    case_candao.append('')
                                    case_candao.append(case_name)
                                    case_candao.append(before)
                                    case_candao.append(setp)
                                    case_candao.append(except_result)
                                    case_candao.append('')
                                    case_candao.append(priority)
                                    case_candao.append('功能测试')
                                    case_candao.append('功能测试阶段')
                                    case_candao.append('正常')
                                    self.writeExcel_candao(self.rowNum, case_candao)
                                    self.rowNum += 1
        return error

    def modify_public_excel(self):
        wb = openpyxl.load_workbook(os.getcwd()+'/src/excel_output/'+self.excel_public_name)
        ws = wb['Sheet']
        font = Font(name=u'宋体', bold=True)
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        fill = PatternFill("solid", fgColor="6495ED")
        for i in range(1, ws.max_column + 1):
            ws.cell(row=1, column=i).font = font
            ws.cell(row=1, column=i).fill = fill
            ws.cell(row=1, column=i).alignment = align
        start_b = 2
        start_c = 2
        start_str_b = ''
        start_str_c = ''
        for i in range(2, ws.max_row + 2):
            if i == 2:
                start_str_b = ws.cell(row=2, column=2).value
                start_str_c = ws.cell(row=2, column=3).value
            if start_str_b != ws.cell(row=i, column=2).value:
                ws.merge_cells('B' + str(start_b) + ':B' + str(i - 1))
                ws['B' + str(start_b)].font = font
                ws['B' + str(start_b)].alignment = align
                start_b = i
                start_str_b = ws.cell(row=i, column=2).value
            if start_str_c != ws.cell(row=i, column=3).value:
                ws.merge_cells('C' + str(start_c) + ':C' + str(i - 1))
                ws['C' + str(start_c)].font = font
                ws['C' + str(start_c)].alignment = align
                start_c = i
                start_str_c = ws.cell(row=i, column=3).value
            ws.cell(row=i, column=4).alignment = align
            ws.cell(row=i, column=5).alignment = align
        wb.save(os.getcwd()+'/src/excel_output/'+self.excel_public_name)
        self.get_zip()

    def get_zip(self):
        zf = zipfile.ZipFile('./src/excel_output/'+self.zip_name, mode='w')
        files = ['./src/excel_output/'+self.excel_public_name,'./src/excel_output/'+self.excel_candao_name]
        for file in files:
            zf.write(file)
        zf.close()

#
# if __name__ == '__main__':
#     filename_xmind = 'D:\\pyproject\\融创-知识库测试用例.xmind'
#     template = 'D:\\pyproject\\caseTemplate.xlsx'
#     obj = xmind_to_xlsx()
#     obj.load_xmind(filename_xmind, template)
#     obj.write_excel(template)
#     obj.modify_public_excel()

