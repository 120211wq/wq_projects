# encoding: utf-8
import openpyxl
from common.function import find_path

def load_data(file_path, sheet):
    '''
    加载表格数据
    :param file: 文件名称
    :param sheet: 表名称
    :return: 将每行单元格的值组装成的list
    '''

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet]
    lists = []
    for i in range(2, ws.max_row + 1):
        row_lists = []
        for j in range(1, ws.max_column + 1):
            value = ws.cell(row=i, column=j).value
            if value is None:
                value = ''
            row_lists.append(value)
        lists.append(row_lists)
    return lists


def get_sheets(file_path):
    wb = openpyxl.load_workbook(file_path)
    return wb.get_sheet_names()


def load_product_data(file_path, sheet):
    '''
    加载表格数据
    :param file_path: 文件名称
    :param sheet: 表名称
    :return: 将每行单元格的值组装成的list
    '''

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet]
    lists = []
    for i in range(2, ws.max_row + 1):
        row_lists = []
        for j in range(1, ws.max_column + 1):
            value = ws.cell(row=i, column=j).value
            if value is None:
                value = ''
            row_lists.append(value)
        lists.append(row_lists)
    return lists


def load_data_by_column(file_path, sheet, name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet]
    lists = []
    for i in range(1, ws.max_column+1):
        if ws.cell(row=1, column=i).value == name:
            for j in range(2, ws.max_row + 1):
                value = ws.cell(row=j, column=i).value
                if value is None:
                    value = ''
                lists.append(value)
            return lists
    else:
        print("找不到:"+ name)
        return None


def load_data_by_row_and_col(file_path, sheet, col_len, row_len=1):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet]
    lists = []
    for i in range(1, col_len + 1):
        col_lists = []
        for j in range(2, row_len + 2):
            value = ws.cell(row=j, column=i).value
            if value is None:
                value = ''
            col_lists.append(value)
        lists.append(col_lists)
    return lists


def write_product_data(file_path, sheet, lists):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet]
        print(len(lists))
        j = 0
        for i in range(2,len(lists)+2):
            ws.cell(row=i, column=1).value = lists[j]["pname"]
            ws.cell(row=i, column=2).value = lists[j]["pid"]
            ws.cell(row=i, column=3).value = lists[j]["pkey"]
            j = j +1
            wb.save(file_path)
        print("保存成功")
    except Exception as msg:
        print("保存失败,错误信息：%s" %msg)


def write_project_data(file_path, sheet, lists):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet]
        print(len(lists))
        j = 0
        for i in range(2,len(lists)+2):
            ws.cell(row=i, column=1).value = lists[j]["pj_name"]
            ws.cell(row=i, column=2).value = lists[j]["pj_id"]
            j = j +1
            wb.save(file_path)
        print("保存成功")
    except Exception as msg:
        print("保存失败,错误信息：%s" %msg)


def clear_product_data(file_path, sheet):
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb[sheet]
        # print(ws.max_row)
        ws.delete_rows(2,ws.max_row)
        wb.save(file_path)
        print("清除数据成功")
        # print(ws.max_row)
    except Exception as msg:
        print("清除失败,错误信息：%s" %msg)


def get_product_name(file_path, sheet_name, loc_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=2).value == loc_name:
            break
    wb.close()
    return ws.cell(row=i, column=1).value


def get_locator_value(file_path, sheet_name, loc_name):
    '''
    :param file_path:
    :param sheet_name:
    :param loc_name:
    :return:
    '''
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=1).value == loc_name:
            break
    wb.close()
    return ws.cell(row=i, column=2).value


def get_column_idx(file_path, sheet_name, column_name):
    '''
    :param file_path:
    :param sheet_name:
    :param column_name:
    :return:
    '''
    wb = openpyxl.load_workbook(file_path)
    ws = wb.get_sheet_by_name(sheet_name)
    for i in range(1, ws.max_column+1):
        if ws.cell(row=1, column=i).value == column_name:
            break
    wb.close()
    return i


def get_key_value(file_path, sheet_name, key):
    '''
    获取键值
    :param file_path:
    :param sheet_name:
    :param key:
    :return: key
    '''
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == key:
            return ws.cell(row=i, column=2).value


def write_value(file_path, sheet_name, id_list, pid):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    id = 0
    ws.cell(row=1, column=1).value = pid
    for i in range(2, len(id_list) + 1):
        ws.cell(row=i, column=1).value = id_list[id]
        id = id +1
    wb.save(file_path)
    print("保存成功")

def load_room_data(file_path, sheet):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet]
    lists = []
    for i in range(2, ws.max_row + 1):
        row_lists = []
        for j in range(1, ws.max_column + 1):
            value = ws.cell(row=i, column=j).value
            if value is None:
                value = ''
            row_lists.append(value)
        lists.append(row_lists)
    return lists

def write_key_value(file_path, sheet_name, key, value):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=1).value == key:
            ws.cell(row=i, column=2).value = value
            wb.save(file_path)


def column_length(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    lists = []
    for i in range(1,ws.max_row+1):
        counter = 0
        for j in range(1, ws.max_column+1):
            if ws.cell(row=i, column=j).value != None:
                counter +=1
        lists.append(counter)
    lists.sort()
    lists.reverse()
    return lists[0]


def row_length(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    lists = []
    for i in range(1,ws.max_column+1):
        counter = 0
        for j in range(1, ws.max_row+1):
            if ws.cell(row=j, column=i).value != None:
                counter +=1
        lists.append(counter)
    lists.sort()
    lists.reverse()
    return lists[0]


def get_key_values(file_path, sheet_name, key):
    '''
    获取键值
    :param file_path:
    :param sheet_name:
    :param key:
    :return: key
    '''
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    lists = []
    for i in range(1, ws.max_row+1):
        if ws.cell(row=i, column=1).value == key:
            for j in range(3,ws.max_column+1):
                lists.append(ws.cell(row=i, column=j).value)
            return lists
    else:
        print("未找到:"+ key)
        print(ws.cell(row=2,column=1).value)
        print(ws.cell(row=2,column=1).value)
        return None

# def write_possible_judge(file_path, sheet_name):
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb[sheet_name]
#     ws.cell(row=1, column=1).value = "test"
#     wb.save(file_path)


# def get_key_values(file_path, sheet_name):
#     '''
#     获取键值
#     :param file_path:
#     :param sheet_name:
#     :param key:
#     :return: key
#     '''
#     wb = openpyxl.load_workbook(file_path)
#     ws = wb[sheet_name]
#     lists = []
#     for i in range(1, ws.max_row+1):
#         lists.append(ws.cell(row=i, column=1).value)
#     return lists


if __name__ == '__main__':
    path = find_path() + '/data' + "/login.xlsx"
    # path = find_path() + '/data' + "/products.xlsx"
    # path = find_path() + '/data' + "/devices.xlsx"
    # path = find_path() + '/data' + "/parking.xlsx"
    # path = find_path() + '/data' + "/projects.xlsx"
    # print(row_length(path, "Sheet1"))
    # print(load_data_by_column(path, "项目信息","项目类型"))
    # write_possible_judge(path, "Sheet1")


