# encoding: utf-8
import openpyxl


def load_analysis_data(file_path, sheet):
    '''
    加载表格数据
    :param file_path: 文件名称
    :param sheet: 表名称n
    :return: 将每行单元格的值组装成的list
    '''

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet]
    lists = []
    for j in range(2, ws.max_column + 1):
        value = ws.cell(row=3, column=j).value
        if value is None:
            value = ''
        lists.append(value)
    return lists
